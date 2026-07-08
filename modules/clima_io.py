"""
Canal de clima (El Niño / La Niña) para as projeções do Copom.

Contexto
--------
No modelo agregado 2024Q2 (mAgregado2024q2_base.mod) o clima entra na curva de
Phillips de preços livres:

    piL_t = ... + climaA - climaB + eps_piL

com

    climaA = média_{j=0,1,2}  (α5·dEL_{t-j} + α6·dLA_{t-j})·climaSq_{t-j}   (alta)
    climaB = média_{j=3,4,5}  (α5·dEL_{t-j} + α6·dLA_{t-j})·climaSq_{t-j}   (baixa)

onde dEL/dLA são dummies El Niño/La Niña e climaSq é a intensidade (RONI²).
α5=0.0012 (El Niño), α6=0.0007 (La Niña).

IMPORTANTE: o termo é um PRODUTO de variáveis nulas no steady state, então o
canal é identicamente ZERO na solução de 1ª ordem do Dynare (ghx/ghu). Alimentar
eps_EL/eps_clima no motor não produz efeito. Como climaA-climaB entra em piL_t de
forma aditiva com coeficiente 1 — idêntico a eps_piL — reproduzimos o impacto do
BCB calculando o forçamento f = climaA - climaB aqui e somando-o ao eps_piL bruto.

Fonte da RONI observada: NOAA CPC
    https://www.cpc.ncep.noaa.gov/data/indices/RONI.ascii.txt
(colunas SEAS YR ANOM; médias móveis de 3 meses). A projeção (previsão CFSv2 que o
BCB cita no RPM) é inserida manualmente pelo usuário. As duas são guardadas por
reunião em data/roni_ref.xlsx (uma aba por reunião).
"""

from typing import Optional
from datetime import date, timedelta
import numpy as np
import pandas as pd
import openpyxl
import requests
from pathlib import Path

# Coeficientes do .mod (aalpha5 = El Niño, aalpha6 = La Niña)
ALPHA5 = 0.0012
ALPHA6 = 0.0007

# Limiares de classificação NOAA/BCB (°C)
LIMIAR_EL = 0.5
LIMIAR_LA = -0.5

RONI_URL = "https://www.cpc.ncep.noaa.gov/data/indices/RONI.ascii.txt"

# Estação móvel de 3 meses → trimestre-calendário
SEAS_TO_Q = {"JFM": 1, "AMJ": 2, "JAS": 3, "OND": 4}

RONI_FILE = Path(__file__).parent.parent / "data" / "roni_ref.xlsx"
RONI_COLS = ["Período", "RONI"]


# ---------------------------------------------------------------------------
# Utilidades de trimestre ("2026Q3")
# ---------------------------------------------------------------------------
def _q_index(q: str) -> int:
    """'2026Q3' → índice inteiro absoluto (ano*4 + (tri-1))."""
    y, t = int(q[:4]), int(q[-1])
    return y * 4 + (t - 1)


def _q_label(i: int) -> str:
    y, t = divmod(i, 4)
    return f"{y}Q{t + 1}"


def shift_quarter(q: str, k: int) -> str:
    """Desloca o trimestre em k posições (k negativo = para trás)."""
    return _q_label(_q_index(q) + k)


# ---------------------------------------------------------------------------
# RONI observada (NOAA CPC)
# ---------------------------------------------------------------------------
def fetch_roni_observada(url: str = RONI_URL, timeout: int = 20) -> dict:
    """
    Baixa a RONI do NOAA CPC e retorna {trimestre: RONI} apenas para os
    trimestres-calendário (JFM=Q1, AMJ=Q2, JAS=Q3, OND=Q4).
    Retorna {} em caso de falha de rede.
    """
    try:
        r = requests.get(url, timeout=timeout)
        r.raise_for_status()
        texto = r.text
    except Exception:
        return {}

    out: dict = {}
    for line in texto.splitlines():
        parts = line.split()
        if len(parts) != 3:
            continue
        seas, yr, anom = parts
        if seas not in SEAS_TO_Q:
            continue
        try:
            y = int(yr)
            a = float(anom)
        except ValueError:
            continue  # cabeçalho ou linha inválida
        out[f"{y}Q{SEAS_TO_Q[seas]}"] = a
    return out


# ---------------------------------------------------------------------------
# RONI → forçamento em preços livres (climaA - climaB)
# ---------------------------------------------------------------------------
def _termo_clima(roni: Optional[float]) -> float:
    """(α5·dEL + α6·dLA)·climaSq, com climaSq = RONI² e dummies por limiar."""
    if roni is None:
        return 0.0
    dEL = 1.0 if roni >= LIMIAR_EL else 0.0
    dLA = 1.0 if roni <= LIMIAR_LA else 0.0
    return (ALPHA5 * dEL + ALPHA6 * dLA) * (roni ** 2)


def _iri_seasons_to_quarters(series: list, last_obs_month: int, year: int,
                             corr: float = 0.0) -> dict:
    """
    Mapeia a lista de previsão do plume IRI (9 estações móveis) para trimestres
    -calendário. Convenção: forecast[i] = estação de 3 meses começando no mês
    (last_obs_month − 1 + i); guardamos só as alinhadas ao trimestre (início em
    jan/abr/jul/out). Aplica `corr` (correção relativa) subtraindo do valor.
    """
    out = {}
    for i, val in enumerate(series):
        if val is None or val == -999:
            continue
        sm = last_obs_month - 1 + i          # mês inicial da estação
        yy = year + (sm - 1) // 12
        mm0 = ((sm - 1) % 12) + 1
        if mm0 in (1, 4, 7, 10):
            q = (mm0 - 1) // 3 + 1
            out[f"{yy}Q{q}"] = round(val - corr, 2)
    return out


def fetch_plume_iri(year: int, month: int, timeout: int = 20) -> Optional[dict]:
    """
    Baixa o plume ENSO do IRI/Columbia e devolve projeções de RONI por trimestre.

    Endpoint: https://ensoforecast.iri.columbia.edu/plumes_json/{year}/{month}

    Retorna dict com:
      'month'          : mês do último observado no plume
      'correcao'       : correção relativa (observado_abs − observado_roni)
      'roni_ensemble'  : {trimestre: valor} da MÉDIA do ensemble multi-modelo
                         (averages.total) — empiricamente reproduz o número do BCB
                         (ex.: OND/2026 = 2.1 na safra de maio; 1.49 na de março)
      'roni_relative'  : {trimestre: RONI} da média dos modelos relativos (curta)
      'roni_cfsv2'     : {trimestre: RONI} = NCEP CFSv2 absoluto − correcao
                         (modelo citado pelo BCB, mas cru; superestima ~0.4)
      'observed_roni'  : {trimestre: RONI observada}
    Retorna None em caso de falha/JSON indisponível.

    Nota: o BCB usa CFSv2 "PDF+Spread corrected"; empiricamente o número
    publicado bate com a média do ensemble (`roni_ensemble`), não com o CFSv2
    cru — a recalibração puxa o CFSv2 agressivo para o consenso multi-modelo.
    """
    url = f"https://ensoforecast.iri.columbia.edu/plumes_json/{year}/{int(month):02d}"
    try:
        r = requests.get(url, timeout=timeout,
                         headers={"User-Agent": "Mozilla/5.0"})
        r.raise_for_status()
        d = r.json()
    except Exception:
        return None

    M = d.get("month")
    if M is None:
        return None
    av = d.get("averages", {}) or {}

    # correção relativa a partir do último ponto observado
    corr = 0.0
    obs, oroni = d.get("observed"), d.get("observed_roni")
    if obs and oroni:
        try:
            corr = float(obs[-1]["data"]) - float(oroni[-1]["data"])
        except Exception:
            corr = 0.0

    cfsv2 = next((m for m in d.get("models", [])
                  if m.get("model") == "NCEP CFSv2"), None)
    cfsv2_abs = cfsv2["data"] if cfsv2 else []

    def _obs_to_q(lst):
        out = {}
        for e in (lst or []):
            mon = e.get("month")
            if mon in SEAS_TO_Q:  # só estações de 3 meses alinhadas ao trimestre
                out[f"{year}Q{SEAS_TO_Q[mon]}"] = round(float(e["data"]), 2)
        return out

    _extrap = extrapolar_proximo_trimestre(av.get("total", []), M, year)

    return {
        "month": M,
        "correcao": round(corr, 2),
        "roni_ensemble": _iri_seasons_to_quarters(av.get("total", []), M, year),
        # trimestre estimado 1 tri além do horizonte do plume (marcar com *)
        "roni_ensemble_extrap": ({_extrap[0]: _extrap[1]} if _extrap else {}),
        "roni_relative": _iri_seasons_to_quarters(av.get("relative", []), M, year),
        "roni_cfsv2": _iri_seasons_to_quarters(cfsv2_abs, M, year, corr=corr),
        "observed_roni": _obs_to_q(oroni),
    }


def fetch_plume_iri_cutoff(cutoff: date, timeout: int = 20,
                           max_recuos: int = 6) -> Optional[dict]:
    """
    Puxa do plume do IRI a safra que estava **disponível até o cutoff** (a
    sexta-feira anterior à reunião do Copom). O IRI publica ~dia 19: se o cutoff
    for antes do dia 19, a safra vigente é a do mês anterior. Recua até
    `max_recuos` meses se a safra não estiver arquivada.

    Retorna o mesmo dict de `fetch_plume_iri` acrescido de 'vintage' ("YYYY-MM").
    """
    y, m = cutoff.year, cutoff.month
    if cutoff.day < 19:          # plume do mês corrente ainda não publicado
        prev = date(y, m, 1) - timedelta(days=1)
        y, m = prev.year, prev.month

    for _ in range(max_recuos):
        p = fetch_plume_iri(y, m, timeout=timeout)
        if p:
            p["vintage"] = f"{y}-{m:02d}"
            return p
        prev = date(y, m, 1) - timedelta(days=1)
        y, m = prev.year, prev.month
    return None


def extrapolar_proximo_trimestre(total: list, last_obs_month: int,
                                 year: int) -> Optional[tuple]:
    """
    Estima o trimestre-calendário **imediatamente após** a cobertura do plume,
    quando ele termina exatamente 1 mês antes (ex.: última estação DJF → próximo
    JFM). Método: supõe março = média(jan, fev), com dez ≈ NDJ, resultando em
    `JFM = (3·DJF − NDJ)/2` (i.e., (3·última − penúltima)/2 das estações móveis).

    Retorna (rótulo_trimestre, valor) ou None se não aplicável.
    """
    valid = [(i, v) for i, v in enumerate(total) if v is not None and v != -999]
    if len(valid) < 2:
        return None
    (i_sec, v_sec), (i_last, v_last) = valid[-2], valid[-1]
    if i_last - i_sec != 1:                      # precisam ser estações consecutivas
        return None
    nxt = (last_obs_month - 1 + i_last) + 1      # início do trimestre seguinte
    mm0 = ((nxt - 1) % 12) + 1
    if mm0 not in (1, 4, 7, 10):                 # só se cair exatamente num trimestre
        return None
    yy = year + (nxt - 1) // 12
    q = (mm0 - 1) // 3 + 1
    return f"{yy}Q{q}", round((3 * v_last - v_sec) / 2.0, 2)


def estimar_ar1_roni(historico: Optional[list] = None,
                     limiar: float = 0.0) -> float:
    """
    Estima o phi de um AR(1) sem intercepto (reversão a 0) na RONI trimestral.
    Usa a série observada do NOAA se `historico` não for fornecido.

    `limiar` condiciona a estimação às transições em que o valor anterior é
    > limiar — útil para capturar a reversão de eventos de El Niño (ex.: 0.5)
    ou fortes (ex.: 1.0), que decaem um pouco mais rápido que a média (0.78 geral;
    ~0.76 para RONI>0.5; ~0.72 para RONI>1.0).
    """
    if historico is None:
        obs = fetch_roni_observada()
        historico = [obs[q] for q in sorted(obs, key=_q_index)]
    x = np.asarray(historico, dtype=float)
    if len(x) < 10:
        return 0.78  # fallback razoável (ENSO trimestral)
    x0, x1 = x[:-1], x[1:]
    if limiar:
        mask = x0 > limiar
        if mask.sum() >= 8:
            x0, x1 = x0[mask], x1[mask]
    denom = float(np.sum(x0 * x0))
    return round(float(np.sum(x0 * x1) / denom), 3) if denom else 0.78


def projetar_reversao_ar1(anchor: float, n: int,
                          phi: Optional[float] = None) -> list:
    """Projeta `n` trimestres de reversão a 0 via AR(1): x_t = phi·x_{t-1}."""
    if phi is None:
        phi = estimar_ar1_roni()
    out, v = [], float(anchor)
    for _ in range(n):
        v *= phi
        out.append(round(v, 2))
    return out


def construir_caminho_roni(known: dict, quarters: list, phi: float) -> tuple:
    """
    Monta o caminho de RONI sobre `quarters`:
      - usa `known` {trimestre: valor} (observado + móveis divulgados/estimados),
        interpolando eventuais buracos internos;
      - do último trimestre conhecido em diante, **decai por AR(1)**: x_t = phi·x_{t-1}
        (reversão a 0). phi=1 → mantém; phi→0 → vai a zero no próximo trimestre.

    Retorna (path {trimestre: valor}, decaidos {set de trimestres da cauda}).
    """
    known = {k: v for k, v in (known or {}).items() if v is not None}
    if not known:
        return {q: 0.0 for q in quarters}, set()
    last_q = max(known, key=_q_index)
    last_i = _q_index(last_q)
    conhecidos = [q for q in quarters if _q_index(q) <= last_i]
    path = dict(interpolar_roni(known, conhecidos))
    decaidos = set()
    v = float(known[last_q])
    for q in sorted((q for q in quarters if _q_index(q) > last_i), key=_q_index):
        v *= phi
        path[q] = round(v, 2)
        decaidos.add(q)
    return path, decaidos


def interpolar_roni(known: dict, quarters: list) -> dict:
    """
    Preenche um caminho de RONI por interpolação linear entre pontos conhecidos.

    `known` = {trimestre: valor} (observado NOAA + âncoras do RPM, ex.: pico 2,1
    no 4º tri e normalização em 2027). `quarters` = lista ordenada de trimestres
    a preencher. Antes da 1ª âncora e depois da última, segura o valor constante.
    Retorna {trimestre: valor} para todos os `quarters`.
    """
    pts = sorted((_q_index(q), v) for q, v in known.items() if v is not None)
    if not pts:
        return {q: None for q in quarters}

    out = {}
    for q in quarters:
        i = _q_index(q)
        if i <= pts[0][0]:
            out[q] = pts[0][1]
        elif i >= pts[-1][0]:
            out[q] = pts[-1][1]
        else:
            for (i0, v0), (i1, v1) in zip(pts, pts[1:]):
                if i0 <= i <= i1:
                    out[q] = v0 + (v1 - v0) * (i - i0) / (i1 - i0)
                    break
    return out


def roni_para_forcamento(roni_by_q: dict, meeting_q: str, n: int) -> list:
    """
    Converte um caminho de RONI no forçamento f[t] = climaA[t] - climaB[t]
    (pp em piL trimestral) para os `n` trimestres do horizonte a partir de
    `meeting_q`.

    Parâmetros
    ----------
    roni_by_q : dict {trimestre: RONI}
        Deve incluir ~5 trimestres de pré-história antes de `meeting_q`
        (os lags -1..-5 de climaA/climaB alcançam o passado).
    meeting_q : str
        Primeiro trimestre do horizonte (ex.: '2026Q3').
    n : int
        Número de trimestres do horizonte.

    Retorna
    -------
    list[float] de comprimento n (forçamento por trimestre; 0 fora de eventos).
    """
    i0 = _q_index(meeting_q)

    def term_at(i: int) -> float:
        return _termo_clima(roni_by_q.get(_q_label(i)))

    f = []
    for t in range(n):
        i = i0 + t
        climaA = sum(term_at(i - j) for j in (0, 1, 2)) / 3.0
        climaB = sum(term_at(i - j) for j in (3, 4, 5)) / 3.0
        f.append(round(climaA - climaB, 6))
    return f


# ---------------------------------------------------------------------------
# Persistência por reunião — data/roni_ref.xlsx (molde de baseline_io.py)
# ---------------------------------------------------------------------------
def list_roni_refs(file_path: str = str(RONI_FILE)) -> list:
    """Nomes das reuniões com RONI de referência salva."""
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        names = wb.sheetnames
        wb.close()
        return names
    except Exception:
        return []


def load_roni_ref(meeting_name: str, file_path: str = str(RONI_FILE)) -> Optional[dict]:
    """
    Carrega a RONI de referência de uma reunião.
    Retorna {'quarters': [...], 'RONI': [...]} ou None se não encontrada.
    """
    try:
        df = pd.read_excel(file_path, sheet_name=meeting_name)
    except Exception:
        return None
    if "Período" not in df.columns or "RONI" not in df.columns:
        return None
    return {
        "quarters": df["Período"].astype(str).tolist(),
        "RONI":     [None if pd.isna(v) else float(v) for v in df["RONI"].tolist()],
    }


def save_roni_ref(meeting_name: str, quarters: list, roni: list,
                  file_path: str = str(RONI_FILE)) -> None:
    """
    Salva o caminho de RONI (observado + projetado) de uma reunião.
    Cria o arquivo se não existir; sobrescreve a aba se já existir; preserva
    as demais abas.
    """
    n = max(len(quarters), len(roni))

    def _pad(lst):
        return list(lst) + [None] * (n - len(lst))

    rows = list(zip(_pad(quarters), _pad(roni)))

    path = Path(file_path)
    if path.exists():
        wb = openpyxl.load_workbook(file_path)
    else:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    if meeting_name in wb.sheetnames:
        del wb[meeting_name]
    ws = wb.create_sheet(meeting_name)

    for c, col in enumerate(RONI_COLS, start=1):
        ws.cell(1, c).value = col
    for r, row in enumerate(rows, start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(r, c).value = val

    wb.save(file_path)
