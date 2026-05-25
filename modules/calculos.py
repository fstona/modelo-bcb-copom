"""
Pós-processamento dos resultados da simulação.  # v2 — bcb_proj support

Replica as fórmulas das colunas J-AC do copom2026_novo.xlsx:
  K-M  : compounding cumulativo dos choques trimestrais (IPCA, Livres, Adm)
  V    : projeção indireta = Livres*(1-peso) + Adm*peso
  AA-AC: projeção final = baseline RPM + choque acumulado
  X    : diferença IPCA_RPM vs projeção indireta arredondada
"""

import numpy as np
import pandas as pd
from typing import Optional


# ---------------------------------------------------------------------------
# Compounding cumulativo — replica PRODUCT(col_range/100+1)-1)*100
# ---------------------------------------------------------------------------

def compound_cumulative(quarterly_series: np.ndarray, window: int = 4) -> np.ndarray:
    """
    Replica exatamente as fórmulas K-M do copom2026_novo.xlsx.

    Padrão das fórmulas Excel (janela de 4 trimestres):
      K2  = PRODUCT(G2:G2)   → 1 trimestre
      K3  = PRODUCT(G2:G3)   → 2 trimestres
      K4  = PRODUCT(G2:G4)   → 3 trimestres
      K5  = PRODUCT(G2:G5)   → 4 trimestres (janela completa)
      K6  = PRODUCT(G3:G6)   → desliza: descarta G2, acrescenta G6
      K7  = PRODUCT(G4:G7)   → ...
      (janela sempre de 4 trimestres a partir do 5º período)

    Parameters
    ----------
    quarterly_series : array de shape (T,)
        Desvios trimestrais (coluna G/H/I do engine output).
    window : int
        Tamanho da janela (padrão = 4 trimestres = 1 ano).

    Returns
    -------
    result : array de shape (T,)
    """
    T = len(quarterly_series)
    factors = 1.0 + quarterly_series / 100.0
    result = np.empty(T)
    for t in range(T):
        if t < window:
            # Janela cresce do início: produto de factors[0..t]
            result[t] = (np.prod(factors[:t + 1]) - 1.0) * 100.0
        else:
            # Janela deslizante de tamanho `window`
            result[t] = (np.prod(factors[t - window + 1: t + 1]) - 1.0) * 100.0
    return result


# ---------------------------------------------------------------------------
# Tabela completa de resultados
# ---------------------------------------------------------------------------

def build_results_table(
    sim_output: np.ndarray,
    rpm_ipca,
    rpm_livres,
    rpm_adm,
    quarters: Optional[list] = None,
    peso_adm: float = 0.25,
    sparse_rpm: bool = False,
) -> pd.DataFrame:
    """
    Constrói a tabela de resultados completa equivalente ao Excel.

    Parameters
    ----------
    sim_output : np.ndarray (T × 9)
        Saída bruta do engine (colunas A-I).
    rpm_ipca, rpm_livres, rpm_adm : list
        Projeções do BCB.
        - sparse_rpm=False (1ª reunião): lista contígua de N períodos a partir do início.
          Posições N+1 em diante recebem NaN.
        - sparse_rpm=True (2ª reunião): lista de exatamente 16 valores com NaN nas
          posições sem dado e valores nos horizontes específicos do Copom.
    sparse_rpm : bool
        Indica se o baseline é esparso (2ª reunião) ou contíguo (1ª reunião).
    """
    T = sim_output.shape[0]

    if quarters is None:
        quarters = [f"Q{t+1}" for t in range(T)]

    col_names = ["it", "expectativa", "delta_e", "ht", "ICbr",
                 "Brent", "IPCA", "Livres", "Adm"]

    df = pd.DataFrame(sim_output, columns=col_names)
    df.insert(0, "Período", (quarters + [f"Q{i+1}" for i in range(T)])[:T])

    # Colunas K-M: compounding
    df["IPCA_cum"]   = compound_cumulative(df["IPCA"].values)
    df["Livres_cum"] = compound_cumulative(df["Livres"].values)
    df["Adm_cum"]    = compound_cumulative(df["Adm"].values)

    if sparse_rpm:
        # Lista já tem 16 posições com NaN onde não há dado
        df["RPM_IPCA"]   = list(rpm_ipca)[:T]
        df["RPM_Livres"] = list(rpm_livres)[:T]
        df["RPM_Adm"]    = list(rpm_adm)[:T]
    else:
        # Lista contígua: preenche do início, NaN no restante
        n = len(rpm_ipca)
        df["RPM_IPCA"]   = (list(rpm_ipca)   + [np.nan] * T)[:T]
        df["RPM_Livres"] = (list(rpm_livres) + [np.nan] * T)[:T]
        df["RPM_Adm"]    = (list(rpm_adm)    + [np.nan] * T)[:T]

    df["Indireto"]       = df["RPM_Livres"] * (1 - peso_adm) + df["RPM_Adm"] * peso_adm
    df["Indireto_arred"] = df["Indireto"].round(1)
    df["Dif_vs_RPM"]     = df["RPM_IPCA"] - df["Indireto_arred"]
    df["Proj_IPCA"]      = df["Indireto"]   + df["IPCA_cum"]
    df["Proj_Livres"]    = df["RPM_Livres"] + df["Livres_cum"]
    df["Proj_Adm"]       = df["RPM_Adm"]    + df["Adm_cum"]

    return df


# ---------------------------------------------------------------------------
# Exportação Excel — replica o layout do copom2026_novo.xlsx
# ---------------------------------------------------------------------------

def build_comparison_table(
    df: pd.DataFrame,
    copom_horizons: list,
    quarters: list,
    prev_name: str,
    curr_name: str,
    prev_proj: Optional[dict] = None,
    prev_extra: Optional[dict] = None,
    bcb_proj: Optional[dict] = None,
    bcb_name: str = "BCB pub.",
) -> tuple:
    """
    Monta a tabela comparativa (equivalente AF2:AJ9 do copom2026_novo.xlsx).

    Colunas: horizontes do Copom + 1 horizonte extra (índice seguinte).
    Linhas:  IPCA, Livres, Adm  ×  {reunião anterior, reunião atual}.

    Parameters
    ----------
    prev_proj : dict ou None
        Dados carregados de baseline_io.load_meeting() — contém 'Proj_IPCA',
        'Proj_Livres', 'Proj_Adm' como listas de 16 valores. Quando fornecido,
        é usado para toda a linha "Anterior" (incluindo o horizonte extra).
        Tem prioridade sobre rpm_col do df e prev_extra.
    prev_extra : dict ou None
        Fallback manual {'IPCA', 'Livres', 'Adm'} para o horizonte extra quando
        prev_proj não está disponível. Ignorado se prev_proj for fornecido.

    Returns
    -------
    (table_df, hr_col) onde hr_col é o rótulo da coluna do horizonte relevante.
    """
    extra_idx = copom_horizons[-1]["index"] + 1
    all_horizons = list(copom_horizons)
    if extra_idx < len(quarters):
        eq = quarters[extra_idx]
        ey, eq_num = int(eq[:4]), int(eq[-1])
        extra_label = f"{ey} (anual)" if eq_num == 4 else eq
        all_horizons.append({"index": extra_idx, "quarter": eq, "label": extra_label})

    # Rótulo de coluna — acrescenta "(HR)" ao horizonte relevante (índice 6)
    col_labels, hr_col = [], None
    for h in all_horizons:
        lbl = h["label"] + " (HR)" if h["index"] == 6 else h["label"]
        col_labels.append(lbl)
        if h["index"] == 6:
            hr_col = lbl

    # prev_proj usa chaves simples "IPCA"/"Livres"/"Adm" (projeções oficiais do BCB)
    vars_config = [
        ("IPCA",   "RPM_IPCA",   "Proj_IPCA",   "IPCA"),
        ("Livres", "RPM_Livres", "Proj_Livres",  "Livres"),
        ("Adm",    "RPM_Adm",    "Proj_Adm",     "Adm"),
    ]

    rows, idx_tuples = [], []
    for var_name, rpm_col, proj_col, prev_key in vars_config:
        prev_row = []
        for h in all_horizons:
            idx = h["index"]
            if prev_proj is not None:
                # Prioridade: projeções oficiais do BCB carregadas do arquivo
                series = prev_proj.get(prev_key, [])
                val = series[idx] if idx < len(series) else np.nan
                prev_row.append(np.nan if (isinstance(val, float) and np.isnan(val)) else val)
            elif idx == extra_idx and prev_extra is not None:
                # Fallback manual apenas para o horizonte extra (sem arquivo)
                prev_row.append(prev_extra.get(prev_key, np.nan))
            elif idx < len(df):
                # Fallback: RPM baseline do df atual
                prev_row.append(df[rpm_col].iloc[idx])
            else:
                prev_row.append(np.nan)
        rows.append(prev_row)
        idx_tuples.append((var_name, prev_name))

        # Projeções publicadas pelo BCB para a reunião atual (quando disponíveis)
        if bcb_proj is not None:
            bcb_row = []
            for h in all_horizons:
                series = bcb_proj.get(prev_key, [])
                val = series[h["index"]] if h["index"] < len(series) else np.nan
                bcb_row.append(np.nan if (isinstance(val, float) and np.isnan(val)) else val)
            rows.append(bcb_row)
            idx_tuples.append((var_name, bcb_name))

        curr_row = [
            df[proj_col].iloc[h["index"]] if h["index"] < len(df) else np.nan
            for h in all_horizons
        ]
        rows.append(curr_row)
        idx_tuples.append((var_name, curr_name))

    multi_idx = pd.MultiIndex.from_tuples(idx_tuples, names=["Variável", "Reunião"])
    table_df = pd.DataFrame(rows, index=multi_idx, columns=col_labels)
    return table_df, hr_col


def export_excel(df: pd.DataFrame, path, sheet_name: str) -> None:
    """
    Exporta a tabela no mesmo layout do copom2026_novo.xlsx.

    `path` pode ser um caminho de arquivo (str) ou um objeto file-like (BytesIO).

    Colunas A-I  : desvios trimestrais brutos
    Coluna J     : rótulo do período
    Colunas K-M  : compounding acumulado (IPCA, Livres, Adm)
    Colunas R-U  : baseline RPM (IPCA, Livres, Adm, Indireto)
    Colunas W-X  : indireto arredondado, diferença vs RPM
    Colunas AA-AC: projeção atualizada
    """
    import openpyxl

    is_file_path = isinstance(path, str)
    if is_file_path:
        try:
            wb = openpyxl.load_workbook(path)
        except FileNotFoundError:
            wb = openpyxl.Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]
    else:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    raw_cols    = ["it", "expectativa", "delta_e", "ht", "ICbr",
                   "Brent", "IPCA", "Livres", "Adm"]
    cum_cols    = ["IPCA_cum", "Livres_cum", "Adm_cum"]
    rpm_cols    = ["RPM_IPCA", "RPM_Livres", "RPM_Adm", "Indireto"]
    extra_cols  = ["Indireto_arred", "Dif_vs_RPM"]
    proj_cols   = ["Proj_IPCA", "Proj_Livres", "Proj_Adm"]

    headers_A = raw_cols                                   # A1-I1
    headers_J = ["Período"]                               # J1
    headers_K = ["IPCA_cum", "Livres_cum", "Adm_cum"]    # K1-M1 (col 11-13)
    headers_R = ["RPM_IPCA", "RPM_Livres", "RPM_Adm", "Indireto",  # R1-U1 (col 18-21)
                 "Indireto_arred", "Dif_vs_RPM"]                   # V1-W1
    headers_AA = proj_cols                                # AA1-AC1 (col 27-29)

    # Linha 1: headers nas colunas corretas
    for i, h in enumerate(headers_A, start=1):
        ws.cell(1, i).value = h
    ws.cell(1, 10).value = "Período"
    for i, h in enumerate(headers_K, start=11):
        ws.cell(1, i).value = h
    for i, h in enumerate(headers_R, start=18):
        ws.cell(1, i).value = h
    for i, h in enumerate(headers_AA, start=27):
        ws.cell(1, i).value = h

    for row_idx, row in df.iterrows():
        r = int(row_idx) + 2

        # A-I: desvios brutos
        for c, col in enumerate(raw_cols, start=1):
            ws.cell(r, c).value = row[col]

        # J: período
        ws.cell(r, 10).value = row["Período"]

        # K-M: compounding acumulado
        for c, col in enumerate(cum_cols, start=11):
            ws.cell(r, c).value = row[col]

        # R-W: baseline RPM + indireto
        for c, col in enumerate(["RPM_IPCA", "RPM_Livres", "RPM_Adm",
                                  "Indireto", "Indireto_arred", "Dif_vs_RPM"], start=18):
            ws.cell(r, c).value = row[col]

        # AA-AC: projeção final
        for c, col in enumerate(proj_cols, start=27):
            ws.cell(r, c).value = row[col]

    wb.save(path if is_file_path else path)
