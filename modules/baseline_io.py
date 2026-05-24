"""
Persistência das projeções oficiais do BCB entre reuniões do Copom.

Arquivo: projecoes_copom.xlsx
Formato: uma sheet por reunião, editável diretamente no Excel.

Colunas: Período | IPCA | Livres | Adm
Valores: projeções OFICIAIS do BCB (acumulado 4 trimestres), mesmo formato
         da aba Baseline do app.

Uso:
  - "Reunião atual"   → preenche automaticamente a aba Baseline (RPM)
  - "Reunião anterior"→ preenche a linha "Anterior" na tabela comparativa

As projeções SIMULADAS geradas pelo app não são salvas neste arquivo.
"""

from typing import Optional
import numpy as np
import pandas as pd
import openpyxl
from pathlib import Path

COLS = ["Período", "IPCA", "Livres", "Adm"]


def list_meetings(file_path: str) -> list:
    """Retorna os nomes das reuniões salvas no arquivo."""
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        names = wb.sheetnames
        wb.close()
        return names
    except Exception:
        return []


def load_meeting(file_path: str, meeting_name: str) -> Optional[dict]:
    """
    Carrega as projeções oficiais do BCB de uma reunião específica.

    Returns
    -------
    dict com chaves 'quarters', 'IPCA', 'Livres', 'Adm' (listas de até 16 valores),
    ou None se a reunião não for encontrada.
    """
    try:
        df = pd.read_excel(file_path, sheet_name=meeting_name)
    except Exception:
        return None

    return {
        "quarters": df["Período"].tolist() if "Período" in df.columns else [],
        "IPCA":     df["IPCA"].tolist()    if "IPCA"    in df.columns else [],
        "Livres":   df["Livres"].tolist()  if "Livres"  in df.columns else [],
        "Adm":      df["Adm"].tolist()     if "Adm"     in df.columns else [],
    }


def save_bcb_projections(
    file_path: str,
    meeting_name: str,
    quarters: list,
    ipca: list,
    livres: list,
    adm: list,
) -> None:
    """
    Salva as projeções oficiais do BCB para uma reunião no arquivo de referência.

    Cria o arquivo se não existir; sobrescreve a sheet se já existir.
    Preserva sheets de outras reuniões.
    """
    n = max(len(quarters), len(ipca), len(livres), len(adm))

    def _pad(lst):
        return list(lst) + [None] * (n - len(lst))

    rows = list(zip(_pad(quarters), _pad(ipca), _pad(livres), _pad(adm)))

    path = Path(file_path)
    if path.exists():
        wb = openpyxl.load_workbook(file_path)
    else:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    if meeting_name in wb.sheetnames:
        del wb[meeting_name]
    ws = wb.create_sheet(meeting_name)

    for c, col in enumerate(COLS, start=1):
        ws.cell(1, c).value = col
    for r, row in enumerate(rows, start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(r, c).value = val

    wb.save(file_path)
